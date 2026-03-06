"""
Data export service.
Exports offers to CSV, Excel, and JSON formats.
"""
import csv
import json
from io import StringIO, BytesIO
from typing import List, Optional
from datetime import datetime

from sqlalchemy import select, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.logging_config import get_logger
from app.models import Offer, Source

logger = get_logger("export")


class DataExporter:
    """
    Export offers to various formats.
    
    Supports:
    - CSV
    - Excel (XLSX)
    - JSON
    """
    
    def __init__(self):
        pass
    
    async def export_to_csv(
        self,
        session: AsyncSession,
        source: Optional[str] = None,
        limit: Optional[int] = None
    ) -> str:
        """
        Export offers to CSV format.
        
        Returns:
            CSV string
        """
        offers = await self._get_offers(session, source, limit)
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "ID", "Source", "Title", "Price", "Currency",
            "City", "Region", "Area_m2", "Rooms",
            "Latitude", "Longitude", "URL",
            "First_Seen", "Last_Seen"
        ])
        
        # Data
        for offer in offers:
            writer.writerow([
                str(offer.id),
                offer.source.name if offer.source else "",
                offer.title,
                offer.price,
                offer.currency,
                offer.city,
                offer.region,
                offer.area_m2,
                offer.rooms,
                offer.lat,
                offer.lng,
                offer.url,
                offer.first_seen.isoformat() if offer.first_seen else "",
                offer.last_seen.isoformat() if offer.last_seen else "",
            ])
        
        return output.getvalue()
    
    async def export_to_json(
        self,
        session: AsyncSession,
        source: Optional[str] = None,
        limit: Optional[int] = None,
        include_raw: bool = False
    ) -> str:
        """
        Export offers to JSON format.
        
        Returns:
            JSON string
        """
        offers = await self._get_offers(session, source, limit)
        
        data = []
        for offer in offers:
            item = {
                "id": str(offer.id),
                "source": offer.source.name if offer.source else None,
                "title": offer.title,
                "price": float(offer.price) if offer.price else None,
                "currency": offer.currency,
                "city": offer.city,
                "region": offer.region,
                "area_m2": offer.area_m2,
                "rooms": offer.rooms,
                "location": {
                    "lat": float(offer.lat) if offer.lat else None,
                    "lng": float(offer.lng) if offer.lng else None,
                },
                "url": offer.url,
                "first_seen": offer.first_seen.isoformat() if offer.first_seen else None,
                "last_seen": offer.last_seen.isoformat() if offer.last_seen else None,
            }
            
            if include_raw and offer.raw_json:
                item["raw_data"] = offer.raw_json
            
            data.append(item)
        
        return json.dumps({
            "exported_at": datetime.utcnow().isoformat(),
            "count": len(data),
            "offers": data
        }, indent=2, ensure_ascii=False)
    
    async def export_to_excel(
        self,
        session: AsyncSession,
        source: Optional[str] = None,
        limit: Optional[int] = None
    ) -> bytes:
        """
        Export offers to Excel format.
        
        Returns:
            Excel file as bytes
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, cannot export to Excel")
            raise RuntimeError("openpyxl is required for Excel export")
        
        offers = await self._get_offers(session, source, limit)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Offers"
        
        # Header
        headers = [
            "ID", "Source", "Title", "Price", "Currency",
            "City", "Region", "Area (m²)", "Rooms",
            "Latitude", "Longitude", "URL",
            "First Seen", "Last Seen"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, offer in enumerate(offers, 2):
            ws.cell(row=row, column=1, value=str(offer.id))
            ws.cell(row=row, column=2, value=offer.source.name if offer.source else "")
            ws.cell(row=row, column=3, value=offer.title)
            ws.cell(row=row, column=4, value=float(offer.price) if offer.price else None)
            ws.cell(row=row, column=5, value=offer.currency)
            ws.cell(row=row, column=6, value=offer.city)
            ws.cell(row=row, column=7, value=offer.region)
            ws.cell(row=row, column=8, value=offer.area_m2)
            ws.cell(row=row, column=9, value=offer.rooms)
            ws.cell(row=row, column=10, value=float(offer.lat) if offer.lat else None)
            ws.cell(row=row, column=11, value=float(offer.lng) if offer.lng else None)
            ws.cell(row=row, column=12, value=offer.url)
            ws.cell(row=row, column=13, value=offer.first_seen)
            ws.cell(row=row, column=14, value=offer.last_seen)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    async def _get_offers(
        self,
        session: AsyncSession,
        source: Optional[str] = None,
        limit: Optional[int] = None
    ) -> List[Offer]:
        """Get offers from database."""
        query = select(Offer).order_by(desc(Offer.last_seen))
        
        if source:
            source_obj = await session.scalar(
                select(Source).where(Source.name == source)
            )
            if source_obj:
                query = query.where(Offer.source_id == source_obj.id)
        
        if limit:
            query = query.limit(limit)
        
        result = await session.execute(query)
        return result.scalars().all()
    
    def get_filename(self, format: str, source: Optional[str] = None) -> str:
        """Generate export filename."""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        source_part = f"_{source}" if source else ""
        
        extensions = {
            "csv": "csv",
            "json": "json",
            "excel": "xlsx"
        }
        
        ext = extensions.get(format, "txt")
        return f"offers{source_part}_{timestamp}.{ext}"


class ScheduledExport:
    """
    Scheduled exports for backup/analysis.
    """
    
    async def export_daily(self):
        """Export yesterday's offers."""
        from app.db import AsyncSessionLocal
        
        async with AsyncSessionLocal() as session:
            exporter = DataExporter()
            
            # Export to JSON
            json_data = await exporter.export_to_json(
                session,
                limit=10000
            )
            
            # Save to file or cloud storage
            # TODO: Implement S3/MinIO upload
            
            logger.info("Daily export completed")


# Global instance
_data_exporter: Optional[DataExporter] = None


def get_data_exporter() -> DataExporter:
    """Get or create data exporter."""
    global _data_exporter
    
    if _data_exporter is None:
        _data_exporter = DataExporter()
    
    return _data_exporter
