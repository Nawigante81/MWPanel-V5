"""
Excel Import/Export Service - Import/Eksport Excel

Masowe operacje na ofertach: import z Excel, eksport do Excel,
szablony importu, walidacja danych.
"""

from datetime import datetime
from typing import Optional, List, Dict, Any, BinaryIO, Tuple
from dataclasses import dataclass, field
from enum import Enum
import io
import uuid

from sqlalchemy.orm import Session

from app.core.logging import get_logger

logger = get_logger(__name__)

# Spróbuj zaimportować openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed, Excel operations will be limited")


class ImportStatus(str, Enum):
    """Status importu"""
    PENDING = "pending"
    VALIDATING = "validating"
    IMPORTING = "importing"
    COMPLETED = "completed"
    FAILED = "failed"
    PARTIAL = "partial"


class ImportErrorType(str, Enum):
    """Typ błędu importu"""
    MISSING_REQUIRED = "missing_required"
    INVALID_FORMAT = "invalid_format"
    INVALID_PRICE = "invalid_price"
    DUPLICATE = "duplicate"
    VALIDATION_ERROR = "validation_error"


@dataclass
class ImportError:
    """Błąd importu"""
    row: int
    column: Optional[str]
    error_type: ImportErrorType
    message: str
    value: Any = None
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'row': self.row,
            'column': self.column,
            'error_type': self.error_type.value,
            'message': self.message,
            'value': self.value,
        }


@dataclass
class ImportResult:
    """Wynik importu"""
    status: ImportStatus
    total_rows: int
    imported_count: int
    error_count: int
    errors: List[ImportError] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    import_id: str = field(default_factory=lambda: str(uuid.uuid4()))
    started_at: datetime = field(default_factory=datetime.utcnow)
    completed_at: Optional[datetime] = None
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'import_id': self.import_id,
            'status': self.status.value,
            'total_rows': self.total_rows,
            'imported_count': self.imported_count,
            'error_count': self.error_count,
            'errors': [e.to_dict() for e in self.errors],
            'warnings': self.warnings,
            'started_at': self.started_at.isoformat(),
            'completed_at': self.completed_at.isoformat() if self.completed_at else None,
        }


@dataclass
class ExportConfig:
    """Konfiguracja eksportu"""
    columns: List[str]
    include_headers: bool = True
    date_format: str = "%Y-%m-%d"
    currency_format: str = "#,##0.00"
    
    # Filtry
    filters: Dict[str, Any] = field(default_factory=dict)
    
    # Formatowanie
    header_style: Optional[Dict[str, Any]] = None
    
    def __post_init__(self):
        if self.header_style is None:
            self.header_style = {
                'font': {'bold': True, 'color': 'FFFFFF'},
                'fill': {'color': '4472C4'},
            }


# Standardowe kolumny dla importu/eksportu ofert
OFFER_COLUMNS = {
    'id': {'header': 'ID', 'required': False, 'type': 'string'},
    'title': {'header': 'Tytuł', 'required': True, 'type': 'string'},
    'description': {'header': 'Opis', 'required': False, 'type': 'string'},
    'property_type': {'header': 'Typ nieruchomości', 'required': True, 'type': 'choice', 'choices': ['mieszkanie', 'dom', 'działka', 'lokal', 'biuro', 'magazyn']},
    'transaction_type': {'header': 'Typ transakcji', 'required': True, 'type': 'choice', 'choices': ['sprzedaż', 'wynajem']},
    'price': {'header': 'Cena', 'required': True, 'type': 'number'},
    'currency': {'header': 'Waluta', 'required': False, 'type': 'string', 'default': 'PLN'},
    'area_sqm': {'header': 'Powierzchnia (m²)', 'required': True, 'type': 'number'},
    'rooms': {'header': 'Liczba pokoi', 'required': False, 'type': 'integer'},
    'floor': {'header': 'Piętro', 'required': False, 'type': 'integer'},
    'total_floors': {'header': 'Liczba pięter', 'required': False, 'type': 'integer'},
    'year_built': {'header': 'Rok budowy', 'required': False, 'type': 'integer'},
    'city': {'header': 'Miasto', 'required': True, 'type': 'string'},
    'district': {'header': 'Dzielnica', 'required': False, 'type': 'string'},
    'street': {'header': 'Ulica', 'required': False, 'type': 'string'},
    'zip_code': {'header': 'Kod pocztowy', 'required': False, 'type': 'string'},
    'lat': {'header': 'Szerokość geo', 'required': False, 'type': 'number'},
    'lng': {'header': 'Długość geo', 'required': False, 'type': 'number'},
    'owner_name': {'header': 'Właściciel', 'required': False, 'type': 'string'},
    'owner_phone': {'header': 'Telefon właściciela', 'required': False, 'type': 'string'},
    'owner_email': {'header': 'Email właściciela', 'required': False, 'type': 'string'},
    'status': {'header': 'Status', 'required': False, 'type': 'choice', 'choices': ['aktywna', 'nieaktywna', 'zarezerwowana', 'sprzedana'], 'default': 'aktywna'},
    'commission_percent': {'header': 'Prowizja (%)', 'required': False, 'type': 'number'},
    'source_url': {'header': 'URL źródła', 'required': False, 'type': 'string'},
    'tags': {'header': 'Tagi', 'required': False, 'type': 'string'},
    'notes': {'header': 'Notatki', 'required': False, 'type': 'string'},
}


class ExcelOperationsService:
    """
    Serwis operacji Excel - import i eksport ofert.
    
    Umożliwia:
    - Import ofert z pliku Excel
    - Eksport ofert do Excel
    - Generowanie szablonów importu
    - Walidację danych przed importem
    """
    
    def __init__(self, db_session: Session):
        self.db = db_session
    
    # ==========================================================================
    # Eksport
    # ==========================================================================
    
    async def export_offers(
        self,
        offers: List[Dict[str, Any]],
        config: Optional[ExportConfig] = None,
    ) -> bytes:
        """
        Eksportuj oferty do Excel.
        
        Args:
            offers: Lista ofert do wyeksportowania
            config: Konfiguracja eksportu
        
        Returns:
            Zawartość pliku Excel jako bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel export")
        
        if config is None:
            config = ExportConfig(columns=list(OFFER_COLUMNS.keys()))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Oferty"
        
        # Nagłówki
        if config.include_headers:
            headers = []
            for col in config.columns:
                col_config = OFFER_COLUMNS.get(col, {})
                headers.append(col_config.get('header', col))
            
            ws.append(headers)
            
            # Styl nagłówków
            header_fill = PatternFill(
                start_color=config.header_style.get('fill', {}).get('color', '4472C4'),
                end_color=config.header_style.get('fill', {}).get('color', '4472C4'),
                fill_type='solid'
            )
            header_font = Font(
                bold=config.header_style.get('font', {}).get('bold', True),
                color=config.header_style.get('font', {}).get('color', 'FFFFFF'),
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dane
        for offer in offers:
            row = []
            for col in config.columns:
                value = offer.get(col)
                
                # Formatowanie specjalne
                if col == 'price' and value is not None:
                    value = float(value)
                elif col == 'tags' and isinstance(value, list):
                    value = ', '.join(value)
                
                row.append(value)
            
            ws.append(row)
        
        # Autoszerokość kolumn
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Zapisz do bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    async def export_template(self) -> bytes:
        """
        Wygeneruj szablon Excel do importu ofert.
        
        Returns:
            Plik szablonu jako bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for template generation")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Szablon Importu"
        
        # Nagłówki
        headers = []
        for col_key, col_config in OFFER_COLUMNS.items():
            header = col_config['header']
            if col_config.get('required'):
                header += " *"
            headers.append(header)
        
        ws.append(headers)
        
        # Styl nagłówków
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        required_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        for idx, cell in enumerate(ws[1], 1):
            col_key = list(OFFER_COLUMNS.keys())[idx - 1]
            col_config = OFFER_COLUMNS[col_key]
            
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Dodaj komentarz z opisem
            comment_text = f"Typ: {col_config['type']}"
            if col_config.get('choices'):
                comment_text += f"\nDozwolone: {', '.join(col_config['choices'])}"
            if col_config.get('default'):
                comment_text += f"\nDomyślnie: {col_config['default']}"
            
            from openpyxl.comments import Comment
            cell.comment = Comment(comment_text, "System")
        
        # Przykładowy wiersz
        example_row = []
        for col_key, col_config in OFFER_COLUMNS.items():
            if col_key == 'title':
                example_row.append("Mieszkanie 3-pokojowe, 65m², Warszawa")
            elif col_key == 'property_type':
                example_row.append("mieszkanie")
            elif col_key == 'transaction_type':
                example_row.append("sprzedaż")
            elif col_key == 'price':
                example_row.append(650000)
            elif col_key == 'area_sqm':
                example_row.append(65)
            elif col_key == 'rooms':
                example_row.append(3)
            elif col_key == 'city':
                example_row.append("Warszawa")
            elif col_key == 'district':
                example_row.append("Mokotów")
            elif col_key == 'status':
                example_row.append("aktywna")
            elif col_key == 'commission_percent':
                example_row.append(2.5)
            else:
                example_row.append("")
        
        ws.append(example_row)
        
        # Arkusz z instrukcjami
        ws_info = wb.create_sheet("Instrukcja")
        ws_info.append(["Instrukcja importu ofert"])
        ws_info.append([])
        ws_info.append(["Pola wymagane (oznaczone *):"])
        
        required_fields = [c['header'] for k, c in OFFER_COLUMNS.items() if c.get('required')]
        ws_info.append([', '.join(required_fields)])
        ws_info.append([])
        ws_info.append(["Typy nieruchomości:"])
        ws_info.append(["mieszkanie, dom, działka, lokal, biuro, magazyn"])
        ws_info.append([])
        ws_info.append(["Typy transakcji:"])
        ws_info.append(["sprzedaż, wynajem"])
        ws_info.append([])
        ws_info.append(["Statusy:"])
        ws_info.append(["aktywna, nieaktywna, zarezerwowana, sprzedana"])
        
        # Formatowanie instrukcji
        ws_info['A1'].font = Font(bold=True, size=14)
        
        # Autoszerokość
        for ws_sheet in [ws, ws_info]:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                ws_sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    # ==========================================================================
    # Import
    # ==========================================================================
    
    async def validate_import(
        self,
        file_content: bytes,
    ) -> ImportResult:
        """
        Waliduj plik Excel przed importem.
        
        Args:
            file_content: Zawartość pliku Excel
        
        Returns:
            ImportResult z błędami walidacji
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel import")
        
        result = ImportResult(
            status=ImportStatus.VALIDATING,
            total_rows=0,
            imported_count=0,
            error_count=0,
        )
        
        try:
            wb = load_workbook(filename=io.BytesIO(file_content))
            ws = wb.active
            
            # Pobierz nagłówki
            headers = [cell.value for cell in ws[1]]
            
            # Mapuj nagłówki na klucze kolumn
            header_to_key = {}
            for col_key, col_config in OFFER_COLUMNS.items():
                header = col_config['header']
                if header in headers:
                    header_to_key[header] = col_key
                elif f"{header} *" in headers:
                    header_to_key[f"{header} *"] = col_key
            
            # Sprawdź wymagane kolumny
            required_columns = [k for k, c in OFFER_COLUMNS.items() if c.get('required')]
            missing_required = []
            for req_col in required_columns:
                req_header = OFFER_COLUMNS[req_col]['header']
                if req_header not in headers and f"{req_header} *" not in headers:
                    missing_required.append(req_header)
            
            if missing_required:
                result.errors.append(ImportError(
                    row=0,
                    column=None,
                    error_type=ImportErrorType.MISSING_REQUIRED,
                    message=f"Brakujące wymagane kolumny: {', '.join(missing_required)}",
                ))
                result.error_count += 1
            
            # Waliduj wiersze
            row_idx = 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_idx += 1
                result.total_rows += 1
                
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        header = headers[idx]
                        col_key = header_to_key.get(header)
                        if col_key:
                            row_data[col_key] = value
                
                # Waliduj wymagane pola
                for req_col in required_columns:
                    if req_col in header_to_key.values():
                        if not row_data.get(req_col):
                            result.errors.append(ImportError(
                                row=row_idx,
                                column=OFFER_COLUMNS[req_col]['header'],
                                error_type=ImportErrorType.MISSING_REQUIRED,
                                message=f"Brak wymaganej wartości: {OFFER_COLUMNS[req_col]['header']}",
                                value=row_data.get(req_col),
                            ))
                            result.error_count += 1
                
                # Waliduj typy
                for col_key, value in row_data.items():
                    col_config = OFFER_COLUMNS.get(col_key)
                    if not col_config or value is None:
                        continue
                    
                    col_type = col_config.get('type')
                    
                    if col_type == 'number':
                        try:
                            float(value)
                        except (ValueError, TypeError):
                            result.errors.append(ImportError(
                                row=row_idx,
                                column=col_config['header'],
                                error_type=ImportErrorType.INVALID_FORMAT,
                                message=f"Nieprawidłowa liczba: {value}",
                                value=value,
                            ))
                            result.error_count += 1
                    
                    elif col_type == 'integer':
                        try:
                            int(value)
                        except (ValueError, TypeError):
                            result.errors.append(ImportError(
                                row=row_idx,
                                column=col_config['header'],
                                error_type=ImportErrorType.INVALID_FORMAT,
                                message=f"Nieprawidłowa liczba całkowita: {value}",
                                value=value,
                            ))
                            result.error_count += 1
                    
                    elif col_type == 'choice':
                        choices = col_config.get('choices', [])
                        if choices and str(value).lower() not in [c.lower() for c in choices]:
                            result.errors.append(ImportError(
                                row=row_idx,
                                column=col_config['header'],
                                error_type=ImportErrorType.INVALID_FORMAT,
                                message=f"Nieprawidłowa wartość. Dozwolone: {', '.join(choices)}",
                                value=value,
                            ))
                            result.error_count += 1
            
            if result.error_count == 0:
                result.status = ImportStatus.COMPLETED
            else:
                result.status = ImportStatus.FAILED
            
            result.completed_at = datetime.utcnow()
            
        except Exception as e:
            result.status = ImportStatus.FAILED
            result.errors.append(ImportError(
                row=0,
                column=None,
                error_type=ImportErrorType.VALIDATION_ERROR,
                message=f"Błąd walidacji: {str(e)}",
            ))
            result.error_count += 1
        
        return result
    
    async def import_offers(
        self,
        file_content: bytes,
        organization_id: Optional[str] = None,
        created_by: Optional[str] = None,
        skip_validation: bool = False,
    ) -> ImportResult:
        """
        Importuj oferty z pliku Excel.
        
        Args:
            file_content: Zawartość pliku Excel
            organization_id: ID organizacji
            created_by: ID użytkownika tworzącego
            skip_validation: Pomiń walidację
        
        Returns:
            ImportResult z wynikiem importu
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel import")
        
        # Najpierw waliduj
        if not skip_validation:
            validation_result = await self.validate_import(file_content)
            if validation_result.error_count > 0:
                return validation_result
        
        result = ImportResult(
            status=ImportStatus.IMPORTING,
            total_rows=0,
            imported_count=0,
            error_count=0,
        )
        
        try:
            wb = load_workbook(filename=io.BytesIO(file_content))
            ws = wb.active
            
            # Pobierz nagłówki
            headers = [cell.value for cell in ws[1]]
            
            # Mapuj nagłówki
            header_to_key = {}
            for col_key, col_config in OFFER_COLUMNS.items():
                header = col_config['header']
                if header in headers:
                    header_to_key[header] = col_key
                elif f"{header} *" in headers:
                    header_to_key[f"{header} *"] = col_key
            
            # Importuj wiersze
            for row in ws.iter_rows(min_row=2, values_only=True):
                result.total_rows += 1
                
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        header = headers[idx]
                        col_key = header_to_key.get(header)
                        if col_key:
                            # Konwertuj typy
                            col_config = OFFER_COLUMNS.get(col_key, {})
                            col_type = col_config.get('type')
                            
                            if col_type == 'number' and value is not None:
                                try:
                                    value = float(value)
                                except:
                                    value = None
                            elif col_type == 'integer' and value is not None:
                                try:
                                    value = int(value)
                                except:
                                    value = None
                            
                            row_data[col_key] = value
                
                # Dodaj domyślne wartości
                for col_key, col_config in OFFER_COLUMNS.items():
                    if col_key not in row_data and col_config.get('default'):
                        row_data[col_key] = col_config['default']
                
                # Dodaj metadane
                row_data['organization_id'] = organization_id
                row_data['created_by'] = created_by
                row_data['imported_at'] = datetime.utcnow()
                
                # Zapisz do bazy (w rzeczywistej implementacji)
                # await self._save_offer_to_db(row_data)
                
                result.imported_count += 1
            
            result.status = ImportStatus.COMPLETED
            result.completed_at = datetime.utcnow()
            
            logger.info(f"Imported {result.imported_count} offers")
            
        except Exception as e:
            result.status = ImportStatus.FAILED
            result.errors.append(ImportError(
                row=0,
                column=None,
                error_type=ImportErrorType.VALIDATION_ERROR,
                message=f"Błąd importu: {str(e)}",
            ))
            result.error_count += 1
        
        return result
    
    async def export_offers_by_query(
        self,
        query_filters: Dict[str, Any],
        config: Optional[ExportConfig] = None,
    ) -> bytes:
        """
        Eksportuj oferty na podstawie zapytania.
        
        Args:
            query_filters: Filtry zapytania
            config: Konfiguracja eksportu
        
        Returns:
            Plik Excel jako bytes
        """
        # W rzeczywistej implementacji: pobierz oferty z bazy
        # offers = await self._query_offers(query_filters)
        offers = []
        
        return await self.export_offers(offers, config)
    
    # ==========================================================================
    # Metody pomocnicze
    # ==========================================================================
    
    async def _save_offer_to_db(self, offer_data: Dict[str, Any]):
        """Zapisz ofertę do bazy"""
        # W rzeczywistej implementacji
        pass
    
    async def _query_offers(
        self,
        filters: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """Pobierz oferty z bazy na podstawie filtrów"""
        # W rzeczywistej implementacji
        return []


# Singleton dla service
def get_excel_service(db_session: Session) -> ExcelOperationsService:
    """Get Excel operations service instance"""
    return ExcelOperationsService(db_session)
